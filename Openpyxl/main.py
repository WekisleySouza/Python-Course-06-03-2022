import openpyxl
from random import uniform

'''pedidos = openpyxl.load_workbook('Openpyxl/pedidos.xlsx')


nome_planilhas = pedidos.sheetnames # Pegar nome das planílhas no arquivo.

planilha1 = pedidos['Página1']

# Pegar célula:
print(planilha1['b4'].value)

# Percorrer coluna:
for celula in planilha1['b']:
    if celula.value != None:
        print(celula.value)

# Alterar valor em planilha. (Não altera na planilha original):
planilha1['b3'].value = 2200

# Adicionando valores:
for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')'''

# Criando nova planilha:
planilha = openpyxl.Workbook()

planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 0)

planilha1 = planilha['Planilha1']
planilha1 = planilha['Planilha2']

